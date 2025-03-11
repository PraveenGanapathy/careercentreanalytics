from flask import Flask, render_template, request, jsonify, redirect, url_for, flash
from flask_wtf import FlaskForm
from flask_wtf.csrf import CSRFProtect
# from wtforms import SelectField, DateField, FloatField, SubmitField
#from wtforms.validators import DataRequired
import openpyxl
from datetime import datetime
import os
from azure.storage.blob import BlobServiceClient
import io
import openpyxl
from dotenv import load_dotenv
import secrets
import os
import logging

load_dotenv()
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

connection_string = os.getenv('AZURE_STORAGE_CONNECTION_STRING')
container_name = os.getenv('AZURE_CONTAINER_NAME')
blob_name = os.getenv('AZURE_BLOB_NAME')

def load_workbook():
    blob_service_client = BlobServiceClient.from_connection_string(connection_string)
    container_client = blob_service_client.get_container_client(container_name)
    blob_client = container_client.get_blob_client(blob_name)
    
    download_stream = blob_client.download_blob()
    file_stream = io.BytesIO()
    download_stream.readinto(file_stream)
    file_stream.seek(0)
    return openpyxl.load_workbook(file_stream, data_only=True)


def save_workbook(wb):
    blob_service_client = BlobServiceClient.from_connection_string(connection_string)
    container_client = blob_service_client.get_container_client(container_name)
    blob_client = container_client.get_blob_client(blob_name)
    
    # Save workbook to a bytes stream
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    # Upload the bytes stream to Azure Storage
    blob_client.upload_blob(file_stream, overwrite=True)


app = Flask(__name__)
app.secret_key = secrets.token_hex(16)
csrf = CSRFProtect(app)

# Get absolute path to the Excel file, let it  be here for future local Dev.
# Modify the WORKBOOK_PATH to ensure proper file location
#WORKBOOK_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "CareerCenterMetrics.xlsx")

# def load_workbook():
#     try:
#         logger.info("Attempting to load workbook from Azure")
#         if not os.path.exists(WORKBOOK_PATH):
#             print(f"Excel file not found at: {WORKBOOK_PATH}")
#             return None
#         return openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
#     except Exception as e:
#         print(f"Error loading workbook: {str(e)}")
#         return None


def get_metrics_and_categories():
    try:
        wb = load_workbook()
        if not wb:
            return {}, {}
            
        sheet = wb["MetricsAndCategories"]
        categories_and_metrics = {}
        metric_groups = {}
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 3 and row[0] and row[1]:
                category, metric = row[0], row[1].strip()
                group = row[2] if len(row) >= 3 and row[2] else "Other"
                
                # Store category-metric relationship
                if category not in categories_and_metrics:
                    categories_and_metrics[category] = []
                categories_and_metrics[category].append(metric)
                
                # Store metric-group relationship
                metric_groups[metric] = group
        
        return categories_and_metrics, metric_groups
    except Exception as e:
        print(f"Error in get_metrics_and_categories: {str(e)}")
        return {}, {}


def get_fiscal_years():
    try:
        wb = load_workbook()
        if not wb:
            return ["24/25"]  # Default fallback
            
        sheet = wb["StagingData"]
        fiscal_years = set()
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                fiscal_years.add(str(row[0]))
        
        return sorted(list(fiscal_years)) if fiscal_years else ["24/25"]
    except Exception as e:
        print(f"Error in get_fiscal_years: {str(e)}")
        return ["24/25"]

@app.route('/')
def dashboard():
    try:
        categories_and_metrics, metric_groups = get_metrics_and_categories()
        current_date = datetime.now().strftime("%m/%d/%Y %I:%M %p")
        fiscal_years = get_fiscal_years()
        return render_template('dashboard.html', 
                             categories=list(categories_and_metrics.keys()),
                             current_date=current_date,
                             fiscal_years=fiscal_years,
                             metric_groups=metric_groups)
    except Exception as e:
        print(f"Error in dashboard route: {str(e)}")
        return "Error loading dashboard", 500


@app.route('/form', methods=['GET', 'POST'])
def form():
    try:
        categories_and_metrics, metric_groups = get_metrics_and_categories()
        fiscal_years = get_fiscal_years()
        quarters = ["Q1", "Q2", "Q3", "Q4"]
        
        if request.method == 'POST':
            # Check if it's an AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                try:
                    # Process form data
                    fiscal_year = request.form.get('fiscal_year')
                    quarter = request.form.get('quarter')
                    start_date = request.form.get('start_date')
                    end_date = request.form.get('end_date')
                    category = request.form.get('category')
                    action = request.form.get('action', 'create')
                    
                    # Validate required fields
                    if not all([fiscal_year, quarter, start_date, end_date, category]):
                        return jsonify({
                            'success': False,
                            'message': 'All fields are required!'
                        })
                    
                    # Load previous data for comparison
                    wb = load_workbook()
                    if not wb:
                        return jsonify({
                            'success': False,
                            'message': 'Could not open the Excel file for reading.'
                        })
                        
                    sheet = wb["StagingData"]
                    
                    # Get previous values for metrics
                    previous_values = {}
                    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                        if (row and len(row) >= 6 and 
                            str(row[0]).strip() == fiscal_year and 
                            str(row[1]).strip() == quarter and 
                            str(row[4]).strip() == category):
                            metric_name = row[5]
                            previous_values[metric_name] = row[6]
                    
                    # Process metrics data
                    metrics_data = []
                    
                    # If updating, first remove existing entries
                    if action == 'update':
                        rows_to_delete = []
                        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                            if (row and len(row) >= 6 and 
                                str(row[0]).strip() == fiscal_year and 
                                str(row[1]).strip() == quarter and 
                                str(row[4]).strip() == category):
                                rows_to_delete.append(row_idx)
                        
                        # Delete rows in reverse order to avoid index shifting
                        for row_idx in sorted(rows_to_delete, reverse=True):
                            sheet.delete_rows(row_idx)
                    
                    # Add new entries
                    next_row = sheet.max_row + 1
                    metrics_added = 0
                    metrics_updated = 0
                    
                    # Get all form fields
                    for key, value in request.form.items():
                        if key.startswith('metrics_') and value.strip():
                            metric_name = key[8:]  # Remove 'metrics_' prefix
                            target_key = f'targets_{metric_name}'
                            target_value = request.form.get(target_key, '')
                            
                            # Convert dates to datetime objects
                            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
                            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
                            
                            # Check if this metric already exists
                            existing_row = None
                            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                                if (row and len(row) >= 6 and 
                                    str(row[0]).strip() == fiscal_year and 
                                    str(row[1]).strip() == quarter and 
                                    str(row[4]).strip() == category and 
                                    str(row[5]).strip() == metric_name):
                                    existing_row = row_idx
                                    break
                            
                            if existing_row:
                                # Update existing row
                                sheet.cell(row=existing_row, column=7).value = float(value)
                                sheet.cell(row=existing_row, column=8).value = float(target_value) if target_value.strip() else None
                                metrics_updated += 1
                            else:
                                # Add new row
                                sheet.cell(row=next_row, column=1).value = fiscal_year
                                sheet.cell(row=next_row, column=2).value = quarter
                                sheet.cell(row=next_row, column=3).value = start_date_obj
                                sheet.cell(row=next_row, column=4).value = end_date_obj
                                sheet.cell(row=next_row, column=5).value = category
                                sheet.cell(row=next_row, column=6).value = metric_name
                                sheet.cell(row=next_row, column=7).value = float(value)
                                sheet.cell(row=next_row, column=8).value = float(target_value) if target_value.strip() else None
                                
                                next_row += 1
                                metrics_added += 1
                    
                    # Save the workbook
                    save_workbook(wb)
                    
                    # Generate update message with previous values
                    update_message = []
                    for key, value in request.form.items():
                        if key.startswith('metrics_'):
                            metric_name = key[8:]  # Remove 'metrics_' prefix
                            previous_value = previous_values.get(metric_name, 'N/A')
                            if not value.strip():
                                update_message.append(f"{metric_name}: Reset from {previous_value} to empty")
                            else:
                                update_message.append(f"{metric_name}: Updated from {previous_value} to {value}")
                    
                    return jsonify({
                        'success': True,
                        'message': 'Changes confirmed: ' + ', '.join(update_message),
                        'metrics_count': metrics_added + metrics_updated
                    })
                    
                except Exception as e:
                    print(f"AJAX error: {str(e)}")
                    return jsonify({
                        'success': False,
                        'message': f'Error: {str(e)}'
                    })
            else:
                # Regular form submission (non-AJAX)
                try:
                    fiscal_year = request.form.get('fiscal_year')
                    quarter = request.form.get('quarter')
                    start_date = request.form.get('start_date')
                    end_date = request.form.get('end_date')
                    category = request.form.get('category')
                    
                    # Validate required fields
                    if not all([fiscal_year, quarter, start_date, end_date, category]):
                        flash('All fields are required!', 'error')
                        return render_template('form.html',
                                            fiscal_years=fiscal_years,
                                            quarters=quarters,
                                            categories=list(categories_and_metrics.keys()),
                                            metrics_by_category=categories_and_metrics)
                    
                    # Process metrics data
                    wb = load_workbook()
                    if wb:
                        sheet = wb["StagingData"]
                        next_row = sheet.max_row + 1
                        metrics_added = 0
                        metrics_updated = 0
                        
                        # Load previous values for metrics
                        previous_values = {}
                        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                            if (row and len(row) >= 6 and 
                                str(row[0]).strip() == fiscal_year and 
                                str(row[1]).strip() == quarter and 
                                str(row[4]).strip() == category):
                                metric_name = row[5]
                                previous_values[metric_name] = row[6]
                        
                        for key, value in request.form.items():
                            if key.startswith('metrics_') and value.strip():
                                metric_name = key[8:]  # Remove 'metrics_' prefix
                                target_key = f'targets_{metric_name}'
                                target_value = request.form.get(target_key, '')
                                
                                # Convert dates to datetime objects
                                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
                                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
                                
                                # Check if this metric already exists
                                existing_row = None
                                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                                    if (row and len(row) >= 6 and 
                                        str(row[0]).strip() == fiscal_year and 
                                        str(row[1]).strip() == quarter and 
                                        str(row[4]).strip() == category and 
                                        str(row[5]).strip() == metric_name):
                                        existing_row = row_idx
                                        break
                                
                                if existing_row:
                                    # Update existing row
                                    sheet.cell(row=existing_row, column=7).value = float(value)
                                    sheet.cell(row=existing_row, column=8).value = float(target_value) if target_value.strip() else None
                                    metrics_updated += 1
                                else:
                                    # Add new row
                                    sheet.cell(row=next_row, column=1).value = fiscal_year
                                    sheet.cell(row=next_row, column=2).value = quarter
                                    sheet.cell(row=next_row, column=3).value = start_date_obj
                                    sheet.cell(row=next_row, column=4).value = end_date_obj
                                    sheet.cell(row=next_row, column=5).value = category
                                    sheet.cell(row=next_row, column=6).value = metric_name
                                    sheet.cell(row=next_row, column=7).value = float(value)
                                    sheet.cell(row=next_row, column=8).value = float(target_value) if target_value.strip() else None
                                    
                                    next_row += 1
                                    metrics_added += 1
                        
                        # Save the workbook
                        save_workbook(wb)
                        update_message = []
                        for key, value in request.form.items():
                            if key.startswith('metrics_'):
                                metric_name = key[8:]  # Remove 'metrics_' prefix
                                previous_value = previous_values.get(metric_name, 'N/A')
                                if not value.strip():
                                    update_message.append(f"{metric_name}: Reset from {previous_value} to empty")
                                else:
                                    update_message.append(f"{metric_name}: Updated from {previous_value} to {value}")
                        
                        flash('Changes confirmed: ' + ', '.join(update_message), 'success')
                    else:
                        flash('Could not open the Excel file for writing.', 'error')
                    
                    return redirect(url_for('dashboard'))
                    
                except Exception as e:
                    flash(f'Error saving data: {str(e)}', 'error')
        
        # GET request or form rendering after POST
        return render_template('form.html',
                            fiscal_years=fiscal_years,
                            quarters=quarters,
                            categories=list(categories_and_metrics.keys()),
                            metrics_by_category=categories_and_metrics)
                            
    except Exception as e:
        print(f"Error in form route: {str(e)}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': False,
                'message': f'Error: {str(e)}'
            })
        else:
            flash(f'An error occurred: {str(e)}', 'error')
            return "Error loading form", 500


@app.route('/get_metrics_data')
def get_metrics_data():
    try:
        wb = load_workbook()
        if not wb:
            return jsonify([])
            
        sheet = wb["StagingData"]
        data = []
        
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            try:
                if not row or len(row) < 8:
                    print(f"Skipping row {row_num}: Insufficient columns")
                    continue

                # Process dates
                start_date = None
                end_date = None
                try:
                    if row[2]:
                        start_date = row[2].strftime('%Y-%m-%d') if isinstance(row[2], datetime) else datetime.strptime(str(row[2]), '%Y-%m-%d').strftime('%Y-%m-%d')
                    if row[3]:
                        end_date = row[3].strftime('%Y-%m-%d') if isinstance(row[3], datetime) else datetime.strptime(str(row[3]), '%Y-%m-%d').strftime('%Y-%m-%d')
                except (ValueError, AttributeError) as e:
                    print(f"Date processing error in row {row_num}: {e}")

                # Process numeric values
                try:
                    value = float(str(row[6]).strip()) if row[6] is not None else 0
                    target = float(str(row[7]).strip()) if row[7] is not None else None
                except (ValueError, TypeError, AttributeError):
                    print(f"Numeric value processing error in row {row_num}")
                    continue

                data.append({
                    'fiscal_year': str(row[0]).strip(),
                    'quarter': str(row[1]).strip(),
                    'start_date': start_date,
                    'end_date': end_date,
                    'category': str(row[4]).strip(),
                    'metric': str(row[5]).strip(),
                    'value': value,
                    'target': target
                })
            except Exception as e:
                print(f"Error processing row {row_num}: {str(e)}")
                continue
                
        print(f"Successfully processed {len(data)} rows of data")
        return jsonify(data)
    except Exception as e:
        print(f"Error in get_metrics_data: {str(e)}")
        return jsonify([])

@app.route('/get_metric_groups')
def get_metric_groups():
    try:
        _, metric_groups = get_metrics_and_categories()
        return jsonify(metric_groups)
    except Exception as e:
        print(f"Error in get_metric_groups: {str(e)}")
        return jsonify({})
@app.route('/check_existing_data')
def check_existing_data():
    fiscal_year = request.args.get('fiscal_year')
    quarter = request.args.get('quarter')
    category = request.args.get('category')
    
    if not all([fiscal_year, quarter, category]):
        return jsonify({'exists': False, 'error': 'Missing parameters'})
    
    try:
        wb = load_workbook()
        if not wb:
            return jsonify({'exists': False})
            
        sheet = wb["StagingData"]
        
        # Check if data exists for these parameters
        exists = False
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if (row and len(row) >= 6 and 
                str(row[0]).strip() == fiscal_year and 
                str(row[1]).strip() == quarter and 
                str(row[4]).strip() == category):
                exists = True
                break
                
        return jsonify({'exists': exists})
    except Exception as e:
        print(f"Error checking existing data: {str(e)}")
        return jsonify({'exists': False, 'error': str(e)})

@app.route('/get_existing_data')
def get_existing_data():
    fiscal_year = request.args.get('fiscal_year')
    quarter = request.args.get('quarter')
    category = request.args.get('category')
    
    if not all([fiscal_year, quarter, category]):
        return jsonify([])
    
    try:
        wb = load_workbook()
        if not wb:
            return jsonify([])
            
        sheet = wb["StagingData"]
        data = []
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if (row and len(row) >= 8 and 
                str(row[0]).strip() == fiscal_year and 
                str(row[1]).strip() == quarter and 
                str(row[4]).strip() == category):
                
                # Process dates
                start_date = None
                end_date = None
                try:
                    if row[2]:
                        start_date = row[2].strftime('%Y-%m-%d') if isinstance(row[2], datetime) else datetime.strptime(str(row[2]), '%Y-%m-%d').strftime('%Y-%m-%d')
                    if row[3]:
                        end_date = row[3].strftime('%Y-%m-%d') if isinstance(row[3], datetime) else datetime.strptime(str(row[3]), '%Y-%m-%d').strftime('%Y-%m-%d')
                except (ValueError, AttributeError) as e:
                    print(f"Date processing error: {e}")

                data.append({
                    'fiscal_year': str(row[0]).strip(),
                    'quarter': str(row[1]).strip(),
                    'start_date': start_date,
                    'end_date': end_date,
                    'category': str(row[4]).strip(),
                    'metric': str(row[5]).strip(),
                    'value': float(str(row[6]).strip()) if row[6] is not None else 0,
                    'target': float(str(row[7]).strip()) if row[7] is not None else None
                })
                
        return jsonify(data)
    except Exception as e:
        print(f"Error getting existing data: {str(e)}")
        return jsonify([])
    

@app.route('/get_metrics_by_category')
def get_metrics_by_category():
    category = request.args.get('category')
    
    if not category:
        return jsonify([])
    
    try:
        categories_and_metrics, metric_groups = get_metrics_and_categories()
        
        if category not in categories_and_metrics:
            return jsonify([])
            
        metrics = categories_and_metrics[category]
        result = []
        
        for metric in metrics:
            group = metric_groups.get(metric, 'Other')
            result.append({
                'name': metric,
                'group': group
            })
            
        return jsonify(result)
    except Exception as e:
        print(f"Error getting metrics by category: {str(e)}")
        return jsonify([])



if __name__ == '__main__':
    app.run(debug=True)
